import os
import re
import json
import time
import requests
import pandas as pd
from pathlib import Path
from dotenv import load_dotenv
from openai import OpenAI

load_dotenv()

# لأن evaluation.py داخل backend/evaluation
# ونريد الرجوع لمجلد backend
BASE_DIR = Path(__file__).resolve().parent.parent

BACKEND_URL = "http://127.0.0.1:8000/chat"

DATASET_JSON_PATH = BASE_DIR / "data" / "phishing_knowledge.json"

MAX_EVAL_SAMPLES = 5

EVAL_DIR = BASE_DIR / "evaluation"

LLM_JUDGE_RESULTS_PATH = EVAL_DIR / "llm_judge_results.csv"
RAGAS_LIKE_RESULTS_PATH = EVAL_DIR / "ragas_like_results.csv"
COMPARISON_RESULTS_PATH = EVAL_DIR / "evaluation_comparison.csv"

client = OpenAI()

# =========================
# Helpers
# =========================

def clean_text_for_evaluation(text):
    text = str(text)
    text = text.replace("\n", " ")
    text = text.replace("\r", " ")
    text = text.replace("\t", " ")
    text = re.sub(r"\s+", " ", text).strip()
    return text[:2500]


def label_to_expected_type(label):
    label = str(label).lower().strip()

    phishing_values = [
        "1", "true", "phishing", "malicious",
        "scam", "suspicious", "fraud", "high"
    ]

    if label in phishing_values:
        return "phishing"

    return "benign"


def label_to_ground_truth(expected_type):
    if expected_type == "phishing":
        return (
            "This content should be classified as phishing or suspicious. "
            "A correct answer should mention clear security risks and advise the user "
            "not to click suspicious links, open unknown attachments, or share sensitive data."
        )

    return (
        "This content should be classified as benign or safe if there are no clear phishing indicators. "
        "A correct answer should avoid exaggerating the risk and may still recommend basic caution."
    )


def build_question_from_text(text):
    return f"""
افحص الرسالة التالية وحدد هل هي Phishing أم Safe.
حلل علامات الخطر إن وجدت، ثم أعطني نصيحة أمنية واضحة.

الرسالة:
{text}
""".strip()


def normalize_json_to_list(raw_data):
    if isinstance(raw_data, list):
        return raw_data

    if isinstance(raw_data, dict):
        for key in ["data", "items", "records"]:
            if key in raw_data and isinstance(raw_data[key], list):
                return raw_data[key]

        return list(raw_data.values())

    raise ValueError("Unsupported JSON structure.")


def extract_text_from_item(item):
    if isinstance(item, str):
        return item

    if not isinstance(item, dict):
        return str(item)

    possible_keys = [
        "text", "content", "message", "email", "url",
        "body", "question", "input", "prompt", "description"
    ]

    for key in possible_keys:
        value = item.get(key)
        if value is not None and str(value).strip():
            return value

    return json.dumps(item, ensure_ascii=False)


def extract_label_from_item(item):
    if not isinstance(item, dict):
        return "unknown"

    possible_keys = [
        "label", "risk", "is_phishing", "category",
        "type", "class", "target", "result"
    ]

    for key in possible_keys:
        value = item.get(key)
        if value is not None and str(value).strip():
            return value

    return "unknown"


def safe_json_parse(content, fallback):
    try:
        content = str(content).strip()

        if content.startswith("```json"):
            content = content.replace("```json", "").replace("```", "").strip()
        elif content.startswith("```"):
            content = content.replace("```", "").strip()

        start = content.find("{")
        end = content.rfind("}")

        if start != -1 and end != -1:
            content = content[start:end + 1]

        return json.loads(content)

    except Exception:
        return fallback


def extract_answer_text(answer):
    if isinstance(answer, dict):
        classification = answer.get("classification", "")
        confidence = answer.get("confidence", "")
        reason = answer.get("reason", "")
        advice = answer.get("advice", "")
        title = answer.get("title", "")
        explanation = answer.get("explanation", "")
        summary = answer.get("short_summary", "")

        red_flags = answer.get("red_flags", [])
        if isinstance(red_flags, list):
            red_flags_text = " | ".join([str(x) for x in red_flags])
        else:
            red_flags_text = str(red_flags)

        parts = [
            f"classification: {classification}",
            f"confidence: {confidence}",
            f"title: {title}",
            f"reason: {reason}",
            f"red_flags: {red_flags_text}",
            f"advice: {advice}",
            f"explanation: {explanation}",
            f"summary: {summary}",
        ]

        return " ".join([p for p in parts if p.strip()])

    if isinstance(answer, str):
        try:
            parsed = json.loads(answer)
            return extract_answer_text(parsed)
        except Exception:
            return answer

    return str(answer)




# =========================
# Excel/CSV Export Helpers
# =========================

def _compact_cell_for_excel(value, max_chars=350):
    if pd.isna(value):
        return ""
    value = str(value)
    value = value.replace("\r", " ").replace("\n", " ").replace("\t", " ")
    value = re.sub(r"\s+", " ", value).strip()
    if len(value) > max_chars:
        return value[:max_chars] + " ..."
    return value


def _make_summary_dataframe(df, csv_path):
    name = Path(csv_path).name.lower()
    if "llm_judge" in name:
        preferred_cols = [
            "id", "expected_type", "mode", "intent", "latency",
            "correctness", "groundedness", "safety_advice", "type_match",
            "completeness", "overall_score", "hallucination_risk", "reason"
        ]
    elif "ragas" in name:
        preferred_cols = [
            "id", "mode", "intent", "latency", "faithfulness",
            "answer_relevancy", "context_usage", "overall_score", "reason"
        ]
    else:
        preferred_cols = list(df.columns)

    summary_cols = [c for c in preferred_cols if c in df.columns]
    if not summary_cols:
        summary_cols = list(df.columns)

    summary_df = df[summary_cols].copy()
    for col in summary_df.columns:
        if summary_df[col].dtype == "object":
            summary_df[col] = summary_df[col].apply(lambda x: _compact_cell_for_excel(x, max_chars=220))
    return summary_df


def _format_excel_sheet(worksheet, df, compact=True):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    header_fill = PatternFill("solid", fgColor="EAF2F8")
    header_font = Font(bold=True, color="1F2937")
    thin_border = Border(
        left=Side(style="thin", color="D9E2EC"),
        right=Side(style="thin", color="D9E2EC"),
        top=Side(style="thin", color="D9E2EC"),
        bottom=Side(style="thin", color="D9E2EC")
    )

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
        cell.border = thin_border

    worksheet.row_dimensions[1].height = 24

    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
            cell.border = thin_border
        worksheet.row_dimensions[row[0].row].height = 22

    text_cols = {"question", "context", "ground_truth", "answer", "reason", "accuracy_focus", "speed", "ease_of_use", "tradeoffs"}
    metric_cols = {
        "latency", "correctness", "groundedness", "safety_advice", "type_match",
        "completeness", "overall_score", "faithfulness", "answer_relevancy",
        "context_usage", "average_score", "average_latency"
    }

    for idx, column_name in enumerate(df.columns, start=1):
        col_letter = get_column_letter(idx)
        col_name = str(column_name)

        if col_name in text_cols:
            width = 28 if compact else 35
        elif col_name in metric_cols:
            width = 14
        elif col_name in {"id", "method", "expected_type", "mode", "intent", "hallucination_risk"}:
            width = 18
        else:
            width = 16

        worksheet.column_dimensions[col_letter].width = width

        if col_name in metric_cols:
            for cell in worksheet[col_letter][1:]:
                cell.number_format = "0.000"
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)


def save_results_files(df, csv_path):
    """
    Saves results safely for all devices:
    1) CSV with utf-8-sig so Arabic opens correctly in Excel.
    2) XLSX with a clean Summary sheet + Full_Details sheet.

    Important: This function does NOT change evaluation scores or logic.
    """
    csv_path = Path(csv_path)
    xlsx_path = csv_path.with_suffix(".xlsx")
    csv_path.parent.mkdir(parents=True, exist_ok=True)

    # Full CSV, Arabic-compatible.
    df.to_csv(csv_path, index=False, encoding="utf-8-sig")

    try:
        summary_df = _make_summary_dataframe(df, csv_path)

        details_df = df.copy()
        for col in details_df.columns:
            if details_df[col].dtype == "object":
                details_df[col] = details_df[col].apply(lambda x: _compact_cell_for_excel(x, max_chars=500))

        with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
            summary_df.to_excel(writer, index=False, sheet_name="Summary")
            details_df.to_excel(writer, index=False, sheet_name="Full_Details")

            _format_excel_sheet(writer.sheets["Summary"], summary_df, compact=True)
            _format_excel_sheet(writer.sheets["Full_Details"], details_df, compact=False)

        print(f"Saved CSV: {csv_path}")
        print(f"Saved Excel: {xlsx_path}")

    except Exception as e:
        print(f"Saved CSV: {csv_path}")
        print("Excel export failed. Install openpyxl with: pip install openpyxl")
        print(f"Excel error: {e}")


# =========================
# Load Dataset
# =========================

def load_evaluation_dataset():
    print("Loading local JSON dataset...")

    if not DATASET_JSON_PATH.exists():
        raise FileNotFoundError(f"Dataset not found: {DATASET_JSON_PATH}")

    with open(DATASET_JSON_PATH, "r", encoding="utf-8") as f:
        raw_data = json.load(f)

    raw_items = normalize_json_to_list(raw_data)

    print("JSON loaded successfully")
    print("Total raw items:", len(raw_items))

    evaluation_dataset = []

    phishing_count = 0
    benign_count = 0
    max_each_class = MAX_EVAL_SAMPLES // 2

    for item in raw_items:
        if len(evaluation_dataset) >= MAX_EVAL_SAMPLES:
            break

        raw_text = extract_text_from_item(item)
        raw_label = extract_label_from_item(item)

        clean_text = clean_text_for_evaluation(raw_text)

        if len(clean_text) < 10:
            continue

        expected_type = label_to_expected_type(raw_label)

        if expected_type == "phishing":
            if phishing_count >= max_each_class:
                continue
            phishing_count += 1
        else:
            if benign_count >= max_each_class:
                continue
            benign_count += 1

        evaluation_dataset.append({
            "id": f"EVAL_{len(evaluation_dataset) + 1:04d}",
            "original_text": clean_text,
            "question": build_question_from_text(clean_text),
            "expected_type": expected_type,
            "ground_truth": label_to_ground_truth(expected_type),
            "contexts": [clean_text]
        })

    print(f"Prepared {len(evaluation_dataset)} evaluation samples.")
    print(f"Phishing samples: {phishing_count}")
    print(f"Benign samples: {benign_count}")

    return evaluation_dataset


# =========================
# Backend Call
# =========================

def call_backend(question):
    start_time = time.time()

    response = requests.post(
        BACKEND_URL,
        data={
            "message": question,
            "chat_history": "[]"
        },
        timeout=180
    )

    latency = time.time() - start_time

    response.raise_for_status()
    data = response.json()

    answer_object = data.get("answer", data)
    answer_text = extract_answer_text(answer_object)

    mode = data.get("mode", "unknown")
    intent = data.get("intent", "unknown")

    retrieved_sources = data.get("retrieved_sources", [])
    contexts = []

    if isinstance(retrieved_sources, list):
        for source in retrieved_sources:
            if isinstance(source, dict):
                source_text = source.get("text", "")
                advice = source.get("advice", "")
                risk = source.get("risk", "")
                contexts.append(
                    f"risk: {risk}\ntext: {source_text}\nadvice: {advice}"
                )

    return {
        "answer_object": answer_object,
        "answer_text": answer_text,
        "mode": mode,
        "intent": intent,
        "contexts": contexts,
        "latency": latency,
        "raw_response": data
    }


# =========================
# LLM-as-Judge
# =========================

def llm_as_judge(question, answer, ground_truth, expected_type, mode, intent):
    prompt = f"""
You are a strict evaluator for a Scam & Phishing RAG chatbot.

Evaluate the chatbot answer from 0 to 1.

Important:
- Expected type is either phishing or benign.
- If expected_type is phishing, the answer should identify risk or suspicious indicators.
- If expected_type is benign, the answer should not exaggerate and should not falsely classify it as phishing.
- Give JSON only. No markdown.

Expected type:
{expected_type}

Question:
{question}

Reference answer:
{ground_truth}

Chatbot answer:
{answer}

Backend mode:
{mode}

Backend intent:
{intent}

Required JSON:
{{
  "correctness": 0.0,
  "groundedness": 0.0,
  "safety_advice": 0.0,
  "type_match": 0.0,
  "completeness": 0.0,
  "overall_score": 0.0,
  "hallucination_risk": "low/medium/high",
  "reason": "شرح مختصر بالعربي"
}}
"""

    response = client.chat.completions.create(
        model="gpt-4o-mini",
        temperature=0,
        messages=[
            {"role": "system", "content": "Return valid JSON only. No markdown."},
            {"role": "user", "content": prompt}
        ]
    )

    content = response.choices[0].message.content.strip()

    fallback = {
        "correctness": 0.0,
        "groundedness": 0.0,
        "safety_advice": 0.0,
        "type_match": 0.0,
        "completeness": 0.0,
        "overall_score": 0.0,
        "hallucination_risk": "high",
        "reason": f"Invalid JSON from judge: {content}"
    }

    return safe_json_parse(content, fallback)


def run_llm_judge_evaluation(evaluation_dataset):
    rows = []

    print("\nStarting LLM-as-judge evaluation...")

    for item in evaluation_dataset:
        print(f"Evaluating {item['id']} with LLM-as-judge")

        backend_result = call_backend(item["question"])

        judge_result = llm_as_judge(
            question=item["question"],
            answer=backend_result["answer_text"],
            ground_truth=item["ground_truth"],
            expected_type=item["expected_type"],
            mode=backend_result["mode"],
            intent=backend_result["intent"]
        )

        rows.append({
            "id": item["id"],
            "expected_type": item["expected_type"],
            "question": item["question"],
            "ground_truth": item["ground_truth"],
            "answer": backend_result["answer_text"],
            "mode": backend_result["mode"],
            "intent": backend_result["intent"],
            "latency": round(backend_result["latency"], 3),
            "correctness": judge_result.get("correctness", 0),
            "groundedness": judge_result.get("groundedness", 0),
            "safety_advice": judge_result.get("safety_advice", 0),
            "type_match": judge_result.get("type_match", 0),
            "completeness": judge_result.get("completeness", 0),
            "overall_score": judge_result.get("overall_score", 0),
            "hallucination_risk": judge_result.get("hallucination_risk", "unknown"),
            "reason": judge_result.get("reason", "")
        })

    df = pd.DataFrame(rows)
    save_results_files(df, LLM_JUDGE_RESULTS_PATH)

    print("\nLLM-as-judge finished.")
    print(df)
    print(f"Saved to: {LLM_JUDGE_RESULTS_PATH}")

    return df


# =========================
# RAGAS-like Evaluation
# =========================

def ragas_like_judge(question, answer, context, ground_truth):
    prompt = f"""
You are evaluating RAG quality for a phishing detection chatbot.

Score from 0 to 1:
1. faithfulness: Is the answer supported by the given context?
2. answer_relevancy: Does the answer address the user question?
3. context_usage: Did the answer use the important information from the context?
4. overall_score: final RAG quality score.

Return JSON only. No markdown.

Question:
{question}

Context:
{context}

Reference answer:
{ground_truth}

Chatbot answer:
{answer}

Required JSON:
{{
  "faithfulness": 0.0,
  "answer_relevancy": 0.0,
  "context_usage": 0.0,
  "overall_score": 0.0,
  "reason": "شرح مختصر بالعربي"
}}
"""

    response = client.chat.completions.create(
        model="gpt-4o-mini",
        temperature=0,
        messages=[
            {"role": "system", "content": "Return valid JSON only. No markdown."},
            {"role": "user", "content": prompt}
        ]
    )

    content = response.choices[0].message.content.strip()

    fallback = {
        "faithfulness": 0.0,
        "answer_relevancy": 0.0,
        "context_usage": 0.0,
        "overall_score": 0.0,
        "reason": f"Invalid JSON from RAG judge: {content}"
    }

    return safe_json_parse(content, fallback)


def run_ragas_like_evaluation(evaluation_dataset):
    rows = []

    print("\nStarting RAGAS-like evaluation...")

    for item in evaluation_dataset:
        print(f"Evaluating {item['id']} with RAGAS-like")

        backend_result = call_backend(item["question"])

        if backend_result["contexts"]:
            context = "\n\n".join(backend_result["contexts"])
        else:
            context = item["original_text"]

        ragas_result = ragas_like_judge(
            question=item["question"],
            answer=backend_result["answer_text"],
            context=context,
            ground_truth=item["ground_truth"]
        )

        rows.append({
            "id": item["id"],
            "question": item["question"],
            "context": context,
            "ground_truth": item["ground_truth"],
            "answer": backend_result["answer_text"],
            "mode": backend_result["mode"],
            "intent": backend_result["intent"],
            "latency": round(backend_result["latency"], 3),
            "faithfulness": ragas_result.get("faithfulness", 0),
            "answer_relevancy": ragas_result.get("answer_relevancy", 0),
            "context_usage": ragas_result.get("context_usage", 0),
            "overall_score": ragas_result.get("overall_score", 0),
            "reason": ragas_result.get("reason", "")
        })

    df = pd.DataFrame(rows)
    save_results_files(df, RAGAS_LIKE_RESULTS_PATH)

    print("\nRAGAS-like evaluation finished.")
    print(df)
    print(f"Saved to: {RAGAS_LIKE_RESULTS_PATH}")

    return df


# =========================
# Comparison
# =========================

def run_comparison(llm_judge_df, ragas_like_df):
    comparison_df = pd.DataFrame([
        {
            "method": "LLM-as-judge",
            "average_score": round(llm_judge_df["overall_score"].mean(), 3),
            "average_latency": round(llm_judge_df["latency"].mean(), 3),
            "accuracy_focus": "يقيس صحة التصنيف phishing/benign، الهلوسة، النصيحة الأمنية، واكتمال الإجابة",
            "speed": "أبطأ لأنه يستخدم LLM evaluator",
            "ease_of_use": "سهل ومرن ومناسب جدًا لمشروع الأمن السيبراني",
            "tradeoffs": "أفضل لتقييم phishing/security، لكنه يعتمد على جودة prompt وتكلفة API"
        },
        {
            "method": "RAGAS-like",
            "average_score": round(ragas_like_df["overall_score"].mean(), 3),
            "average_latency": round(ragas_like_df["latency"].mean(), 3),
            "accuracy_focus": "يقيس faithfulness وanswer relevancy وcontext usage",
            "speed": "أبطأ أيضًا لأنه يستخدم LLM evaluator",
            "ease_of_use": "أسهل من RAGAS الحقيقي لأنه لا يحتاج datasets/ragas",
            "tradeoffs": "ليس RAGAS رسمي، لكنه يعطي مقارنة قريبة بدون مشاكل تثبيت"
        }
    ])

    save_results_files(comparison_df, COMPARISON_RESULTS_PATH)

    print("\nFinal comparison:")
    print(comparison_df)
    print(f"Saved to: {COMPARISON_RESULTS_PATH}")

    return comparison_df


# =========================
# Main
# =========================

def main():
    os.makedirs(EVAL_DIR, exist_ok=True)

    evaluation_dataset = load_evaluation_dataset()

    llm_judge_df = run_llm_judge_evaluation(evaluation_dataset)

    ragas_like_df = run_ragas_like_evaluation(evaluation_dataset)

    run_comparison(llm_judge_df, ragas_like_df)

    print("\nEvaluation completed successfully.")


if __name__ == "__main__":
    main()