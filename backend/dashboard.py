import streamlit as st
import pandas as pd
import json
import re
from pathlib import Path
from collections import Counter
from io import BytesIO

st.set_page_config(
    page_title="Scam & Phishing Dashboard",
    page_icon="🛡️",
    layout="wide"
)

LOG_PATH = Path("logs/chat_history.jsonl")
AB_LOG_PATH = Path("logs/ab_testing_results.jsonl")



# =========================
# Helper Functions
# =========================

def clean_excel_value(value):
    """
    Remove hidden control characters that Excel/openpyxl refuses to write.
    This fixes IllegalCharacterError caused by characters like \x01 inside logs.
    """
    if value is None:
        return ""

    if isinstance(value, (list, dict)):
        value = json.dumps(value, ensure_ascii=False)
    else:
        value = str(value)

    # Excel does not allow these control characters inside worksheet cells.
    value = re.sub(r"[\x00-\x08\x0B-\x0C\x0E-\x1F]", " ", value)
    value = value.replace("\ufffe", " ").replace("\uffff", " ")
    value = re.sub(r"[ \t]+", " ", value)
    return value.strip()


def dataframe_to_excel_bytes(df, sheet_name="Sheet1"):
    """
    Convert a DataFrame to an Excel .xlsx file in memory.
    Fixes Arabic display AND removes illegal hidden characters from logs.
    """
    output = BytesIO()
    safe_sheet_name = clean_excel_value(sheet_name)[:31] or "Sheet1"

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        export_df = df.copy()

        # Clean every cell before exporting to Excel.
        for col in export_df.columns:
            export_df[col] = export_df[col].apply(clean_excel_value)

        export_df.to_excel(writer, index=False, sheet_name=safe_sheet_name)

        worksheet = writer.sheets[safe_sheet_name]
        worksheet.freeze_panes = "A2"

        from openpyxl.styles import Font, PatternFill, Alignment

        header_fill = PatternFill("solid", fgColor="EAF2F8")

        for cell in worksheet[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(vertical="center", wrap_text=True)

        text_heavy_columns = {
            "user_message", "question", "context", "ground_truth", "reason", "advice",
            "answer", "answer_text", "red_flags", "bot_answer", "raw_answer"
        }

        for column_cells in worksheet.columns:
            column_letter = column_cells[0].column_letter
            header = str(column_cells[0].value or "")
            header_lower = header.lower()

            max_length = len(header)
            for cell in column_cells[1:]:
                value = str(cell.value) if cell.value is not None else ""
                max_length = max(max_length, min(len(value), 80))

            if header_lower in text_heavy_columns:
                worksheet.column_dimensions[column_letter].width = 45
                for cell in column_cells:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
            else:
                worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 30)

        for row in worksheet.iter_rows(min_row=2):
            worksheet.row_dimensions[row[0].row].height = 32

    output.seek(0)
    return output.getvalue()

def parse_bot_answer(value):
    if not isinstance(value, str):
        return {}

    try:
        return json.loads(value)
    except Exception:
        return {
            "classification": "Parsing Error",
            "confidence": 0,
            "reason": "JSON parsing failed",
            "red_flags": ["JSON parsing failed"],
            "advice": "",
            "retrieval_top_score": 0,
            "retrieval_avg_score": 0,
            "answer_mode": "ERROR"
        }


def normalize_flag(flag):
    flag = str(flag).lower()

    if "رابط" in flag or "link" in flag or "url" in flag or "http" in flag:
        return "Suspicious Link"

    if "otp" in flag or "رمز التحقق" in flag or "verification code" in flag:
        return "OTP / Verification Code Request"

    if "كلمة المرور" in flag or "password" in flag or "login" in flag:
        return "Password / Login Theft"

    if "هوية" in flag or "بطاقة" in flag or "identity" in flag or "id" in flag:
        return "Sensitive Data Request"

    if "جائزة" in flag or "ربحت" in flag or "prize" in flag or "winner" in flag:
        return "Fake Prize / Reward"

    if "استعجال" in flag or "فوراً" in flag or "immediately" in flag or "urgent" in flag or "24" in flag:
        return "Urgency / Pressure"

    if "بنك" in flag or "bank" in flag or "paypal" in flag or "حساب" in flag:
        return "Fake Banking / Account Alert"

    if "ملف" in flag or "download" in flag or "تنزيل" in flag:
        return "Suspicious File Download"

    return "Other Red Flag"


def detect_pattern(user_message, label, red_flags):
    text = str(user_message).lower()
    flags_text = " ".join(red_flags).lower() if isinstance(red_flags, list) else ""
    combined = text + " " + flags_text
    label = str(label).lower()

    if label == "general":
        return "General / Out of Scope"

    if "otp" in combined or "رمز التحقق" in combined:
        return "OTP Theft Scam"

    if "paypal" in combined or "bank" in combined or "بنك" in combined or "حساب" in combined:
        return "Fake Banking / Account Scam"

    if "جائزة" in combined or "ربحت" in combined or "prize" in combined:
        return "Fake Prize Scam"

    if "password" in combined or "كلمة المرور" in combined or "login" in combined:
        return "Credential Theft"

    if "http" in combined or "رابط" in combined or "link" in combined:
        return "Suspicious Link Scam"

    if "urgent" in combined or "فوراً" in combined or "immediately" in combined or "24 hours" in combined:
        return "Urgency / Fear Pressure"

    if label in ["phishing", "scam", "suspicious"]:
        return "Other Scam / Phishing"

    if label == "safe":
        return "Safe / Normal"

    return "Unknown Pattern"


def load_logs():
    rows = []

    if not LOG_PATH.exists():
        return pd.DataFrame()

    with open(LOG_PATH, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()

            if not line:
                continue

            try:
                row = json.loads(line)
                parsed = parse_bot_answer(row.get("bot_answer", ""))

                label = parsed.get("classification", "General")
                confidence = parsed.get("confidence", 0)
                red_flags = parsed.get("red_flags", [])

                if not isinstance(red_flags, list):
                    red_flags = []

                row["label"] = label
                row["confidence"] = confidence
                row["risk_score"] = round(float(confidence) * 100, 2)
                row["reason"] = parsed.get("reason", "")
                row["advice"] = parsed.get("advice", "")
                row["red_flags"] = red_flags
                row["retrieval_top_score"] = parsed.get("retrieval_top_score", 0)
                row["retrieval_avg_score"] = parsed.get("retrieval_avg_score", 0)
                row["mode"] = parsed.get("answer_mode", row.get("mode", "GENERAL"))

                row["pattern"] = detect_pattern(
                    row.get("user_message", ""),
                    row["label"],
                    row["red_flags"]
                )

                rows.append(row)

            except Exception:
                continue

    return pd.DataFrame(rows)




# =========================
# A/B Testing Helper Functions
# =========================

def parse_ab_answer(value):
    if isinstance(value, dict):
        return value
    if not isinstance(value, str):
        return {}
    try:
        return json.loads(value)
    except Exception:
        return {}


def load_ab_logs():
    rows = []

    if not AB_LOG_PATH.exists():
        return pd.DataFrame()

    with open(AB_LOG_PATH, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue

            try:
                row = json.loads(line)
                answer = parse_ab_answer(row.get("answer", {}))

                row["classification"] = row.get("classification") or answer.get("classification", "")
                row["confidence"] = float(row.get("confidence") or answer.get("confidence", 0) or 0)
                row["risk_score"] = round(row["confidence"] * 100, 2)
                row["reason"] = answer.get("reason", "")
                row["advice"] = answer.get("advice", "")
                row["red_flags"] = answer.get("red_flags", [])

                if isinstance(row.get("answer"), dict):
                    row["answer_text"] = json.dumps(row.get("answer"), ensure_ascii=False)
                else:
                    row["answer_text"] = str(row.get("answer", ""))

                rows.append(row)
            except Exception:
                continue

    return pd.DataFrame(rows)

# =========================
# Load Data
# =========================

df = load_logs()
ab_df = load_ab_logs()

st.title("🛡️ Scam & Phishing RAG Chatbot Dashboard")
st.caption(
    "Security monitoring dashboard for phishing detection, RAG/LLM analysis, red flags, scam patterns, and high-risk events."
)

if df.empty:
    st.warning("No logs found. Make sure this file exists: logs/chat_history.jsonl")
    st.stop()

df["timestamp"] = pd.to_datetime(df["timestamp"], errors="coerce")

if "latency" not in df.columns:
    df["latency"] = 0

if "error" not in df.columns:
    df["error"] = ""


# =========================
# Sidebar Filters
# =========================

st.sidebar.header("🔎 Filters")

labels = ["All"] + sorted(df["label"].dropna().astype(str).unique().tolist())
selected_label = st.sidebar.selectbox("Risk Label", labels)

modes = ["All"] + sorted(df["mode"].dropna().astype(str).unique().tolist())
selected_mode = st.sidebar.selectbox("Mode", modes)

patterns = ["All"] + sorted(df["pattern"].dropna().astype(str).unique().tolist())
selected_pattern = st.sidebar.selectbox("Scam Pattern", patterns)

min_score = st.sidebar.slider("Minimum Risk Score", 0, 100, 0)

st.sidebar.divider()
st.sidebar.subheader("🧪 A/B Testing Filters")
if not ab_df.empty:
    ab_models = ["All"] + sorted(ab_df.get("model_used", pd.Series(dtype=str)).dropna().astype(str).unique().tolist())
    selected_ab_model = st.sidebar.selectbox("A/B Model", ab_models)

    ab_variants = ["All"] + sorted(ab_df.get("ab_variant", pd.Series(dtype=str)).dropna().astype(str).unique().tolist())
    selected_ab_variant = st.sidebar.selectbox("A/B Variant", ab_variants)
else:
    selected_ab_model = "All"
    selected_ab_variant = "All"


filtered = df.copy()

if selected_label != "All":
    filtered = filtered[filtered["label"].astype(str) == selected_label]

if selected_mode != "All":
    filtered = filtered[filtered["mode"].astype(str) == selected_mode]

if selected_pattern != "All":
    filtered = filtered[filtered["pattern"].astype(str) == selected_pattern]

filtered = filtered[filtered["risk_score"] >= min_score]


# =========================
# Metrics
# =========================

total_requests = len(filtered)

phishing_alerts = filtered[
    filtered["label"].astype(str).str.lower().str.contains("phishing|scam|malicious|suspicious", na=False)
].shape[0]

safe_messages = filtered[
    filtered["label"].astype(str).str.lower().str.contains("safe", na=False)
].shape[0]

general_questions = filtered[
    filtered["label"].astype(str).str.lower().str.contains("general", na=False)
].shape[0]

high_risk_events = filtered[filtered["risk_score"] >= 80].shape[0]

errors = filtered[
    filtered["label"].astype(str).str.lower().str.contains("error", na=False)
].shape[0]

avg_latency = round(filtered["latency"].mean(), 2) if total_requests > 0 else 0


st.subheader("📌 Overview")

col1, col2, col3 = st.columns(3)
col4, col5, col6 = st.columns(3)

col1.metric("Total Requests", total_requests)
col2.metric("Phishing / Suspicious Alerts", phishing_alerts)
col3.metric("Safe Messages", safe_messages)

col4.metric("High Risk Events", high_risk_events)
col5.metric("General Questions", general_questions)
col6.metric("Errors", errors)

if high_risk_events > 0:
    st.error(f"⚠️ High Risk Alert: {high_risk_events} high-risk phishing/scam events detected.")

st.divider()



# =========================
# A/B Testing Admin Section
# =========================

st.subheader("🧪 A/B Testing - Admin Model Comparison")
st.caption(
    "This section is for the admin only. It shows which model answered each request, "
    "the A/B variant, the user question, model answer, latency, confidence, and retrieval score."
)

if ab_df.empty:
    st.info("No A/B testing logs found yet. Expected file: logs/ab_testing_results.jsonl")
else:
    ab_df["timestamp"] = pd.to_datetime(ab_df["timestamp"], errors="coerce")

    filtered_ab = ab_df.copy()

    if selected_ab_model != "All" and "model_used" in filtered_ab.columns:
        filtered_ab = filtered_ab[filtered_ab["model_used"].astype(str) == selected_ab_model]

    if selected_ab_variant != "All" and "ab_variant" in filtered_ab.columns:
        filtered_ab = filtered_ab[filtered_ab["ab_variant"].astype(str) == selected_ab_variant]

    ab_total = len(filtered_ab)
    ab_avg_latency = round(filtered_ab["latency"].mean(), 2) if ab_total > 0 and "latency" in filtered_ab.columns else 0
    ab_avg_confidence = round(filtered_ab["confidence"].mean(), 2) if ab_total > 0 and "confidence" in filtered_ab.columns else 0
    ab_high_risk = filtered_ab[filtered_ab.get("risk_score", 0) >= 80].shape[0] if ab_total > 0 else 0

    ab_col1, ab_col2, ab_col3, ab_col4 = st.columns(4)
    ab_col1.metric("A/B Requests", ab_total)
    ab_col2.metric("Avg Latency", f"{ab_avg_latency}s")
    ab_col3.metric("Avg Confidence", ab_avg_confidence)
    ab_col4.metric("High Risk A/B Events", ab_high_risk)

    ab_left, ab_right = st.columns(2)

    with ab_left:
        st.markdown("### Requests per Model")
        if "model_used" in filtered_ab.columns and not filtered_ab.empty:
            st.bar_chart(filtered_ab["model_used"].astype(str).value_counts())
        else:
            st.info("No model data available.")

    with ab_right:
        st.markdown("### Average Latency per Model")
        if "model_used" in filtered_ab.columns and "latency" in filtered_ab.columns and not filtered_ab.empty:
            st.bar_chart(filtered_ab.groupby("model_used")["latency"].mean())
        else:
            st.info("No latency data available.")

    st.markdown("### 📋 Question + Answer by Model")

    ab_show_cols = [
        "timestamp",
        "user_message",
        "model_used",
        "ab_variant",
        "mode",
        "classification",
        "risk_score",
        "confidence",
        "latency",
        "top_score",
        "reason",
        "advice",
        "answer_text"
    ]
    ab_show_cols = [c for c in ab_show_cols if c in filtered_ab.columns]

    st.dataframe(
        filtered_ab[ab_show_cols].sort_values("timestamp", ascending=False),
        use_container_width=True
    )

    st.markdown("### 🔎 Inspect One A/B Record")
    if not filtered_ab.empty:
        inspect_df = filtered_ab.sort_values("timestamp", ascending=False).reset_index(drop=True)
        options = [
            f"{i} | {str(row.get('timestamp', ''))[:19]} | {row.get('ab_variant', '')} | {row.get('model_used', '')} | {str(row.get('user_message', ''))[:60]}"
            for i, row in inspect_df.iterrows()
        ]
        selected_record = st.selectbox("Choose a record", options)
        selected_index = int(selected_record.split(" | ")[0])
        selected_row = inspect_df.iloc[selected_index]

        q_col, a_col = st.columns(2)
        with q_col:
            st.markdown("#### User Question")
            st.write(selected_row.get("user_message", ""))
            st.markdown("#### Metadata")
            st.json({
                "model_used": selected_row.get("model_used", ""),
                "ab_variant": selected_row.get("ab_variant", ""),
                "mode": selected_row.get("mode", ""),
                "latency": selected_row.get("latency", 0),
                "top_score": selected_row.get("top_score", 0),
                "classification": selected_row.get("classification", ""),
                "confidence": selected_row.get("confidence", 0),
            })
        with a_col:
            st.markdown("#### Model Answer")
            st.write(selected_row.get("reason", ""))
            st.markdown("#### Advice")
            st.write(selected_row.get("advice", ""))

    ab_excel = dataframe_to_excel_bytes(filtered_ab, sheet_name="AB_Testing")
    st.download_button(
        label="⬇️ Download A/B Testing Logs as Excel",
        data=ab_excel,
        file_name="ab_testing_results.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True
    )

st.divider()

# =========================
# Charts
# =========================

st.subheader("📊 Main Distributions")

left, right = st.columns(2)

with left:
    st.markdown("### Risk Label Distribution")
    risk_counts = filtered["label"].astype(str).value_counts()
    st.bar_chart(risk_counts)

with right:
    st.markdown("### Mode Distribution")
    mode_counts = filtered["mode"].astype(str).value_counts()
    st.bar_chart(mode_counts)


st.subheader("🧠 Most Common Scam Patterns")

pattern_counts = filtered["pattern"].astype(str).value_counts()

if pattern_counts.empty:
    st.info("No scam patterns found yet.")
else:
    st.bar_chart(pattern_counts)


# =========================
# Top Red Flags
# =========================

st.subheader("🚩 Top Red Flags")

all_red_flags = []

for item in filtered["red_flags"]:
    if isinstance(item, list):
        normalized_flags = [normalize_flag(flag) for flag in item]
        all_red_flags.extend(normalized_flags)

if len(all_red_flags) == 0:
    st.info("No red flags found yet.")
else:
    red_flag_counts = Counter(all_red_flags)

    red_flags_df = pd.DataFrame(
        red_flag_counts.items(),
        columns=["Red Flag Category", "Count"]
    ).sort_values("Count", ascending=False)

    left, right = st.columns([1, 1])

    with left:
        st.dataframe(red_flags_df, use_container_width=True)

    with right:
        st.bar_chart(red_flags_df.set_index("Red Flag Category")["Count"])


# =========================
# Time Charts
# =========================

st.divider()

st.subheader("📈 Risk Score Over Time")

risk_df = filtered.dropna(subset=["timestamp"]).sort_values("timestamp")

if not risk_df.empty:
    st.line_chart(risk_df.set_index("timestamp")["risk_score"])
else:
    st.info("No timestamp data available.")


st.subheader("🔍 Retrieval Top Score Over Time")

if not risk_df.empty and "retrieval_top_score" in risk_df.columns:
    st.line_chart(risk_df.set_index("timestamp")["retrieval_top_score"])
else:
    st.info("No retrieval score data available.")


# =========================
# High Risk Events
# =========================

st.divider()

st.subheader("🔥 High Risk Events")

high_risk_df = filtered[filtered["risk_score"] >= 80].copy()

if high_risk_df.empty:
    st.success("No high-risk events found.")
else:
    high_risk_cols = [
        "timestamp",
        "user_message",
        "label",
        "risk_score",
        "mode",
        "pattern",
        "reason",
        "advice"
    ]

    high_risk_cols = [c for c in high_risk_cols if c in high_risk_df.columns]

    st.dataframe(
        high_risk_df[high_risk_cols].sort_values("timestamp", ascending=False),
        use_container_width=True
    )

# =========================
# Static File & Image Risk Detection
# =========================

st.divider()
st.subheader("🧩 File & Image Static Risk Detection")
st.caption(
    "This section documents and monitors the new static file/image risk layer. "
    "It checks risky extensions, double extensions, archives, macros, PDF JavaScript, embedded objects, SVG scripts, and QR/link phishing indicators."
)

static_risk_catalog = [
    {"Risk Signal": "invoice.pdf.exe", "Category": "Double Extension / Executable Disguise", "Why It Matters": "Looks like a PDF, but the real final extension is executable."},
    {"Risk Signal": "photo.jpg.zip", "Category": "Archive Disguised as Image", "Why It Matters": "Looks like an image, but it is actually a compressed archive."},
    {"Risk Signal": "document.docm", "Category": "Macro-Enabled Office File", "Why It Matters": "DOCM files can contain VBA macros."},
    {"Risk Signal": "file.js", "Category": "Script File", "Why It Matters": "JavaScript files can execute commands when opened."},
    {"Risk Signal": "update.dmg", "Category": "macOS Installer Risk", "Why It Matters": "DMG files can install applications on macOS."},
    {"Risk Signal": "payment.scr", "Category": "Executable Screensaver Risk", "Why It Matters": "SCR files are executable and commonly abused in malware delivery."},
    {"Risk Signal": "archive.zip", "Category": "Compressed Archive Risk", "Why It Matters": "ZIP files may hide risky files inside."},
    {"Risk Signal": "PDF with JavaScript", "Category": "PDF JavaScript / Auto Action", "Why It Matters": "PDF files can include JavaScript or automatic actions."},
    {"Risk Signal": "DOCX with macros or many links", "Category": "Office Macro / External Links", "Why It Matters": "Office documents may contain macros or external URLs."},
    {"Risk Signal": "Image with QR code or suspicious link", "Category": "QR / Link Phishing in Image", "Why It Matters": "Images can carry QR codes or visible phishing URLs."},
]

catalog_df = pd.DataFrame(static_risk_catalog)
st.dataframe(catalog_df, use_container_width=True)


def detect_static_file_risk_category_dashboard(value):
    value = str(value).lower()

    if any(k in value for k in ["double extension", "امتداد مزدوج", "pdf.exe", "jpg.exe", "docx.exe"]):
        return "Double Extension / Executable Disguise"
    if any(k in value for k in ["dangerous extension", "امتداد ملف تنفيذي", "executable", ".exe", ".scr", ".bat", ".cmd", ".js", ".vbs", ".ps1"]):
        return "Dangerous Executable / Script Extension"
    if any(k in value for k in ["macro", "macro-enabled", "ماكرو", "ماكروز", "vbaproject", ".docm", ".xlsm", ".pptm"]):
        return "Macro-Enabled Office File"
    if any(k in value for k in ["archive", "compressed", "zip", "rar", "7z", "ملف مضغوط"]):
        return "Compressed Archive Risk"
    if any(k in value for k in ["pdf", "javascript", "openaction", "embeddedfile", "launch", "richmedia", "إجراءات تلقائية"]):
        return "PDF JavaScript / Embedded Object Risk"
    if any(k in value for k in ["external links", "targetmode", "روابط خارجية", "many links"]):
        return "Office External Links Risk"
    if any(k in value for k in ["svg", "script", "javascript:", "image/svg"]):
        return "Image/SVG Script Risk"
    if any(k in value for k in ["qr", "qrcode", "qr code", "رمز qr", "باركود"]):
        return "QR / Link Phishing in Image"

    return None


static_risk_rows = []

for _, row in filtered.iterrows():
    values_to_check = []

    for col in ["user_message", "reason", "advice", "mode", "pattern"]:
        if col in row and pd.notna(row[col]):
            values_to_check.append(str(row[col]))

    flags = row.get("red_flags", [])
    if isinstance(flags, list):
        values_to_check.extend([str(flag) for flag in flags])

    categories = []
    for value in values_to_check:
        category = detect_static_file_risk_category_dashboard(value)
        if category:
            categories.append(category)

    categories = list(dict.fromkeys(categories))

    if categories:
        static_risk_rows.append({
            "timestamp": row.get("timestamp", ""),
            "user_message": row.get("user_message", ""),
            "label": row.get("label", ""),
            "risk_score": row.get("risk_score", 0),
            "mode": row.get("mode", ""),
            "static_file_risk_categories": ", ".join(categories),
            "reason": row.get("reason", ""),
            "advice": row.get("advice", ""),
        })

static_risk_df = pd.DataFrame(static_risk_rows)

c_static_1, c_static_2, c_static_3 = st.columns(3)
c_static_1.metric("Supported Static Signals", len(static_risk_catalog))
c_static_2.metric("Detected Static Risk Events", len(static_risk_df))
c_static_3.metric("File/Image Risk Layer", "Enabled")

if static_risk_df.empty:
    st.info("No static file/image risk events found in current logs yet. Upload a risky sample name/type to see it here.")
else:
    all_categories = []
    for cats in static_risk_df["static_file_risk_categories"]:
        all_categories.extend([c.strip() for c in str(cats).split(",") if c.strip()])

    if all_categories:
        counts_df = pd.DataFrame(
            Counter(all_categories).items(),
            columns=["Static Risk Category", "Count"]
        ).sort_values("Count", ascending=False)

        left_static, right_static = st.columns([1, 1])
        with left_static:
            st.markdown("### Detected Static File/Image Risks")
            st.dataframe(counts_df, use_container_width=True)
        with right_static:
            st.markdown("### Static Risk Distribution")
            st.bar_chart(counts_df.set_index("Static Risk Category")["Count"])

    st.markdown("### Recent Static File/Image Risk Events")
    st.dataframe(
        static_risk_df.sort_values("timestamp", ascending=False),
        use_container_width=True
    )

# =========================
# Full Logs
# =========================

st.subheader("📄 Full Chat Logs")

show_cols = [
    "timestamp",
    "user_message",
    "label",
    "risk_score",
    "mode",
    "pattern",
    "retrieval_top_score",
    "retrieval_avg_score",
    "reason",
    "advice"
]

show_cols = [c for c in show_cols if c in filtered.columns]

st.dataframe(
    filtered[show_cols].sort_values("timestamp", ascending=False),
    use_container_width=True
)


# =========================
# Download
# =========================

# =========================
# Download Logs
# =========================

st.divider()
st.subheader("⬇️ Download Logs (Excel Only)")

logs_excel = dataframe_to_excel_bytes(filtered, sheet_name="Chat_Logs")

st.download_button(
    label="⬇️ Download Filtered Logs as Excel",
    data=logs_excel,
    file_name="chat_dashboard_logs.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    use_container_width=True
)


# =========================
# Evaluation Center
# =========================

import subprocess
import sys

st.divider()
st.subheader("🧪 Evaluation Center")

st.write(
    "Run LLM-as-Judge and RAGAS-like evaluation directly from the dashboard, "
    "then preview and download the generated result files."
)

EVAL_DIR = Path("evaluation")
LLM_RESULTS = EVAL_DIR / "llm_judge_results.csv"
RAGAS_RESULTS = EVAL_DIR / "ragas_like_results.csv"
COMPARISON_RESULTS = EVAL_DIR / "evaluation_comparison.csv"


def read_csv_if_exists(path):
    if path.exists():
        try:
            return pd.read_csv(path)
        except Exception as e:
            st.error(f"Could not read {path.name}: {e}")
            return pd.DataFrame()
    return pd.DataFrame()


col_eval_1, col_eval_2 = st.columns([1, 2])

with col_eval_1:
    run_eval = st.button("🚀 Run Evaluation Now", use_container_width=True)

with col_eval_2:
    st.info(
        "Make sure the backend is running on http://127.0.0.1:8000 before pressing this button."
    )


if run_eval:
    with st.spinner("Running evaluation... this may take a few minutes."):
        try:
            result = subprocess.run(
                [sys.executable, "evaluation/evaluation.py"],
                capture_output=True,
                text=True,
                timeout=1800
            )

            if result.returncode == 0:
                st.success("✅ Evaluation completed successfully!")
                with st.expander("Show Evaluation Terminal Output"):
                    st.text(result.stdout)
            else:
                st.error("❌ Evaluation failed.")
                with st.expander("Show Error Output"):
                    st.text(result.stderr)

        except subprocess.TimeoutExpired:
            st.error("❌ Evaluation stopped because it took too long.")
        except Exception as e:
            st.error(f"❌ Error running evaluation: {e}")


st.markdown("### 📊 Evaluation Results Preview")

llm_df = read_csv_if_exists(LLM_RESULTS)
ragas_df = read_csv_if_exists(RAGAS_RESULTS)
comparison_df = read_csv_if_exists(COMPARISON_RESULTS)

if not comparison_df.empty:
    st.markdown("#### 🔁 Evaluation Methods Comparison")
    st.dataframe(comparison_df, use_container_width=True)

    if "average_score" in comparison_df.columns:
        st.bar_chart(comparison_df.set_index("method")["average_score"])

if not llm_df.empty:
    st.markdown("#### 🧠 LLM-as-Judge Results")

    metric_cols = st.columns(4)

    avg_overall = round(llm_df["overall_score"].mean(), 3) if "overall_score" in llm_df.columns else 0
    avg_correctness = round(llm_df["correctness"].mean(), 3) if "correctness" in llm_df.columns else 0
    avg_groundedness = round(llm_df["groundedness"].mean(), 3) if "groundedness" in llm_df.columns else 0
    avg_latency_eval = round(llm_df["latency"].mean(), 3) if "latency" in llm_df.columns else 0

    metric_cols[0].metric("Avg Overall", avg_overall)
    metric_cols[1].metric("Avg Correctness", avg_correctness)
    metric_cols[2].metric("Avg Groundedness", avg_groundedness)
    metric_cols[3].metric("Avg Latency", avg_latency_eval)

    preview_cols = [
        "id",
        "expected_type",
        "mode",
        "intent",
        "correctness",
        "groundedness",
        "safety_advice",
        "type_match",
        "completeness",
        "overall_score",
        "hallucination_risk",
        "reason"
    ]

    preview_cols = [c for c in preview_cols if c in llm_df.columns]
    st.dataframe(llm_df[preview_cols], use_container_width=True)

else:
    st.warning("LLM-as-Judge results not found yet. Run evaluation first.")


if not ragas_df.empty:
    st.markdown("#### 📚 RAGAS-like Results")

    metric_cols = st.columns(4)

    avg_faithfulness = round(ragas_df["faithfulness"].mean(), 3) if "faithfulness" in ragas_df.columns else 0
    avg_relevancy = round(ragas_df["answer_relevancy"].mean(), 3) if "answer_relevancy" in ragas_df.columns else 0
    avg_context_usage = round(ragas_df["context_usage"].mean(), 3) if "context_usage" in ragas_df.columns else 0
    avg_ragas_score = round(ragas_df["overall_score"].mean(), 3) if "overall_score" in ragas_df.columns else 0

    metric_cols[0].metric("Avg Faithfulness", avg_faithfulness)
    metric_cols[1].metric("Avg Relevancy", avg_relevancy)
    metric_cols[2].metric("Avg Context Usage", avg_context_usage)
    metric_cols[3].metric("Avg Overall", avg_ragas_score)

    preview_cols = [
        "id",
        "mode",
        "intent",
        "faithfulness",
        "answer_relevancy",
        "context_usage",
        "overall_score",
        "reason"
    ]

    preview_cols = [c for c in preview_cols if c in ragas_df.columns]
    st.dataframe(ragas_df[preview_cols], use_container_width=True)

else:
    st.warning("RAGAS-like results not found yet. Run evaluation first.")


st.markdown("### 📥 Download Evaluation Reports (Excel Only)")
st.caption("Use Excel files (.xlsx) for clean Arabic text and better formatting across devices.")


def download_excel_button(csv_path, label):
    """
    Download the Excel version generated by evaluation.py.
    The CSV is intentionally hidden from the dashboard because Excel may display Arabic incorrectly in CSV files.
    """
    xlsx_path = Path(csv_path).with_suffix(".xlsx")

    if xlsx_path.exists():
        with open(xlsx_path, "rb") as f:
            st.download_button(
                label=label,
                data=f,
                file_name=xlsx_path.name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
    else:
        st.button(
            f"{label} — Excel file not available yet",
            disabled=True,
            use_container_width=True
        )
        st.caption(f"Run evaluation first to generate: {xlsx_path}")


d1, d2, d3 = st.columns(3)

with d1:
    download_excel_button(LLM_RESULTS, "⬇️ Download LLM-as-Judge Excel")

with d2:
    download_excel_button(RAGAS_RESULTS, "⬇️ Download RAGAS-like Excel")

with d3:
    download_excel_button(COMPARISON_RESULTS, "⬇️ Download Comparison Excel")
